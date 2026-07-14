"""
DOST-Davao Project Mapper

Streamlit app that ingests a DOST-Davao project Excel workbook, cleans and
consolidates it across divisions, and renders an interactive map with
filterable KPIs and per-project detail views.
"""

import pandas as pd
import streamlit as st
import random
import folium
from folium.features import DivIcon
from folium.plugins import MarkerCluster
from streamlit_folium import st_folium
import io
import os
import re
import requests
import time
import tempfile
import openpyxl
try:
    from PIL import Image, ImageDraw, ImageFont
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.chrome.service import Service
    from webdriver_manager.chrome import ChromeDriverManager
    EXPORT_AVAILABLE = True
except ImportError:
    EXPORT_AVAILABLE = False

# ==========================================
# 1. CONSTANTS
# ==========================================
# --- Canonical internal schema -------------------------------------------
# Every ingestion path (legacy tracker sheets OR the new print-style cost
# list) is normalized down to these exact column names before anything
# else in the app touches the data. Nothing downstream should ever
# reference a raw source-workbook header again.
GROUP_KEY_COL = "Project Title"
AMOUNT_ORIGINAL_COL = "Amount (Original)"
AMOUNT_REVISED_COL = "Amount (Revised)"
EFFECTIVE_FUNDING_COL = "Amount (Effective)"
DATE_APPROVED_COL = "Date of Approval"
DATE_END_COL = "Date of Completion/Terminated"

OUTPUT_SCHEMA = [
    "Division", "No.", "Project Title", "Name Abbreviation", "Location", "Beneficiaries",
    "Implementing Agency", "Name of Proponent", "DOST Coordinator",
    AMOUNT_ORIGINAL_COL, AMOUNT_REVISED_COL,
    DATE_APPROVED_COL, DATE_END_COL,
    "Ongoing", "Completed", "Terminated",
    "Remarks", "Lat", "Long", "Source Sheet",
]

# (canonical_field, [keyword fragments to match against a lowercased,
# whitespace-collapsed header]). A header is assigned to the FIRST field
# in this list whose keywords it matches and that hasn't already been
# claimed by an earlier column in the same sheet.
FIELD_ALIASES = [
    ("Project ID",                    ["project id"]),
    ("Project Title",                 ["project title", "title of the project", "project name"]),
    ("Name Abbreviation",             ["abbreviation", "abbrev", "acronym"]),
    ("Division",                      ["division", "program"]),
    ("DOST Coordinator",              ["dost xi project coordinator", "dost coordinator", "project coordinator"]),
    ("Name of Proponent",             ["name of proponent", "proponent", "project leader", "name of project leader"]),
    ("Implementing Agency",           ["implementing agency", "implementing partner", "agency", "partner"]),
    ("Location",                      ["location", "site"]),
    ("Beneficiaries",                 ["beneficiaries", "beneficiary"]),
    ("Date of Approval",              ["date of approval", "approval date"]),
    ("Date of Completion/Terminated", ["date of completion", "completion/terminated", "terminated"]),
    ("Project status",                ["project status"]),
    ("Remarks",                       ["remarks", "notes"]),
    ("Coordinates",                   ["coordinates", "coordinate", "gps"]),
    ("Lat",                           ["latitude", "lat"]),
    ("Long",                          ["longitude", "long", "lng"]),
    (AMOUNT_REVISED_COL,              ["revised", "adjusted"]),
    (AMOUNT_ORIGINAL_COL,             ["original", "amount of funding", "funding provided", "project cost", "amount"]),
    ("No.",                           ["no.", "no "]),
]

# Non-project rows to skip when a division is delimited by section headers
# (e.g. "LGIA sub-total", signature blocks) rather than by separate sheets.
STOP_MARKERS = ("sub-total", "subtotal", "grand total", "prepared by", "reviewed by", "approved by")

# Raw section/sheet division text -> canonical division code used
# everywhere downstream (colors, filters, KPIs).
DIVISION_NORMALIZATION = {
    "LOCAL GIA PROGRAM": "LGIA", "LGIA": "LGIA",
    "CEST PROGRAM": "CEST", "CEST": "CEST",
    "SMART AND SUSTAINABLE PROGRAM": "SSCP", "SSCP": "SSCP",
}

# Sheets scanned in "Legacy Tracker" ingestion mode (others are ignored).
# "Cost List / Any Template" mode instead scans every sheet and keeps
# whichever ones have a recognizable header row.
DATA_SHEETS = ["CEST", "LGIA", "SSCP"]

# Accepted TRUE/FALSE representations found in a raw "Project status" cell.
BOOL_MAP = {
    "TRUE": True, "FALSE": False,
    "True": True, "False": False,
    "1": True, "0": False,
    "1.0": True, "0.0": False,
}

STATUS_LABELS = ["Ongoing", "Completed", "Terminated"]

# Map styling.
DAVAO_CENTER = [7.06750309148034, 125.60060334232874]
DAVAO_ZOOM_START = 13
DIVISION_COLORS = {"CEST": "#2980b9", "LGIA": "#27ae60", "SSCP": "#8e44ad"}
DEFAULT_MARKER_COLOR = "#7f8c8d"
CLICK_TOLERANCE_DEGREES = 0.0001

MISSING_VALUE_TOKENS = {"nan", "nat", "none", "natype"}

# Local LLM API Configuration
OLLAMA_MODEL = "llama3.2:3b"

# ==========================================
# 2. DATA PROCESSING PIPELINE
# ==========================================
# This engine is template-agnostic: it detects the header row, maps
# whatever headers exist to FIELD_ALIASES, figures out per-sheet whether
# projects are laid out as multi-row status blocks ("merged_status" - the
# original tracker format), section-delimited single rows ("section_headers"
# - the new print-style cost list, where a bare row like "CEST PROGRAM"
# marks a division boundary instead of a separate sheet), or already flat
# one-row-per-project ("flat"). All three converge on OUTPUT_SCHEMA so
# nothing downstream needs to know or care which shape the source file was.

INGESTION_MODES = {
    "Division-Based Template": "legacy",
    "General Template": "auto",
}
<<<<<<< HEAD

=======
>>>>>>> 377714d600e0645afde87abb5929f0b11189bc5a

def clean_header_text(v):
    return re.sub(r"\s+", " ", str(v)).strip().lower() if v is not None else ""


def map_headers(header_cells):
    """Returns dict: canonical_field -> original column header text.
    Each column is assigned to the FIRST alias group it matches that hasn't
    already been claimed by an earlier column (so e.g. two 'cost' columns
    become Original then Revised, in the order they appear)."""
    mapping = {}
    claimed_fields = set()
    for raw in header_cells:
        text = clean_header_text(raw)
        if not text:
            continue
        for field, keywords in FIELD_ALIASES:
            if field in claimed_fields:
                continue
            if any(kw in text for kw in keywords):
                mapping[field] = raw
                claimed_fields.add(field)
                break
    return mapping


def find_header_row(ws, max_scan=25):
    """Picks the row (within the first max_scan rows) with the most recognizable
    field matches -- this is template-agnostic, no fixed header text required."""
    best_row, best_score = None, 0
    for row in ws.iter_rows(min_row=1, max_row=max_scan):
        values = [c.value for c in row]
        mapping = map_headers(values)
        if "Project Title" in mapping and len(mapping) > best_score:
            best_row, best_score = row[0].row, len(mapping)
    return best_row


def detect_structure(df, mapping):
    if "Project status" in mapping:
        return "merged_status"

    title_col = mapping.get("Project Title")
    first_col = df.columns[0]
    if title_col is None:
        return "flat"

    has_title = df[title_col].notna() & (df[title_col].astype(str).str.strip() != "")
    first_filled = df[first_col].notna() & (df[first_col].astype(str).str.strip() != "")
    section_like_rows = (~has_title) & first_filled

    if section_like_rows.sum() >= 1:
        return "section_headers"
    return "flat"


def populate_coordinates_from_temp(unified_df):
    if "Lat" not in unified_df.columns:
        unified_df["Lat"] = pd.NA
    if "Long" not in unified_df.columns:
        unified_df["Long"] = pd.NA

    if "_TEMP_COORDINATES_RAW" in unified_df.columns:
        coords_raw = unified_df["_TEMP_COORDINATES_RAW"].astype(str)
        split_coords = coords_raw.str.split(",", n=1, expand=True)
        unified_df["Lat"] = split_coords[0].str.strip()
        unified_df["Long"] = split_coords[1].str.strip() if split_coords.shape[1] > 1 else pd.NA
        unified_df.loc[coords_raw.isna(), ["Lat", "Long"]] = pd.NA
    elif "_TEMP_LAT_RAW" in unified_df.columns and "_TEMP_LONG_RAW" in unified_df.columns:
        unified_df["Lat"] = unified_df["_TEMP_LAT_RAW"]
        unified_df["Long"] = unified_df["_TEMP_LONG_RAW"]

    return unified_df


def _drop_temp_coord_cols(unified):
    return unified.drop(columns=[c for c in ["_TEMP_COORDINATES_RAW", "_TEMP_LAT_RAW", "_TEMP_LONG_RAW"] if c in unified.columns])


def parse_merged_status(df, mapping, sheet_name):
    """The original tracker layout: each project spans several rows, only
    the first row of a block has the group key populated, and status is a
    TRUE/FALSE column paired with a status-label column."""
    status_col = mapping.get("Project status")
    if not status_col or status_col not in df.columns:
        return pd.DataFrame(columns=OUTPUT_SCHEMA)

    cols = list(df.columns)
    status_idx = cols.index(status_col)
    if status_idx + 1 >= len(cols):
        return pd.DataFrame(columns=OUTPUT_SCHEMA)
    label_col = cols[status_idx + 1]

    group_key_col = mapping.get("Project Title") or df.columns[0]
    is_new_group = df[group_key_col].notna() & (df[group_key_col].astype(str).str.strip() != "")
    df = df.copy()
    df["_group_id"] = is_new_group.cumsum()

    shared_cols = [c for c in cols if c not in (status_col, label_col)]
    df[shared_cols] = df.groupby("_group_id")[shared_cols].ffill()
    df["_status_bool"] = df[status_col].map(BOOL_MAP)

    processed_records = []
    for gid, g in df.groupby("_group_id"):
        first_row = g.iloc[0]
        record = {}

        for label in STATUS_LABELS:
            record[label] = 0

        for field in OUTPUT_SCHEMA:
            if field == "No.":
                record[field] = pd.NA
            elif field == "Division":
                record[field] = first_row.get(mapping["Division"]) if mapping.get("Division") else sheet_name
            elif field in STATUS_LABELS:
                continue
            elif field in ("Lat", "Long"):
                continue
            elif field == "Source Sheet":
                record[field] = sheet_name
            else:
                original_col_name = mapping.get(field)
                record[field] = first_row.get(original_col_name) if original_col_name and original_col_name in first_row.index else pd.NA

        for _, row in g.iterrows():
            label = str(row[label_col]).strip()
            if label in STATUS_LABELS and row["_status_bool"] is True:
                record[label] = 1

        if "Coordinates" in mapping:
            record["_TEMP_COORDINATES_RAW"] = first_row.get(mapping["Coordinates"])
        if "Lat" in mapping:
            record["_TEMP_LAT_RAW"] = first_row.get(mapping["Lat"])
        if "Long" in mapping:
            record["_TEMP_LONG_RAW"] = first_row.get(mapping["Long"])

        processed_records.append(record)

    if not processed_records:
        return pd.DataFrame(columns=OUTPUT_SCHEMA)

    unified = pd.DataFrame(processed_records)
    for col in OUTPUT_SCHEMA:
        if col not in unified.columns:
            unified[col] = pd.NA

    unified = populate_coordinates_from_temp(unified)
    unified = _drop_temp_coord_cols(unified)
    return unified[OUTPUT_SCHEMA]


def parse_section_headers(df, mapping, sheet_name):
    """The new print-style layout: no separate sheet per division. Instead
    a bare row (title blank, first column filled, e.g. 'CEST PROGRAM')
    marks the start of a division block; subtotal/signature rows are
    skipped via STOP_MARKERS."""
    title_col = mapping.get("Project Title")
    if not title_col or title_col not in df.columns:
        return pd.DataFrame(columns=OUTPUT_SCHEMA)

    first_col = df.columns[0]
    processed_records = []
    current_division = sheet_name

    for _, row in df.iterrows():
        row_text = " ".join(str(v) for v in row.values if pd.notna(v)).lower()
        if any(marker in row_text for marker in STOP_MARKERS):
            continue

        title_val = row.get(title_col)
        has_title = isinstance(title_val, str) and title_val.strip() != ""
        first_val = row.get(first_col)
        first_str = str(first_val).strip() if pd.notna(first_val) else ""

        if not has_title and first_str != "":
            current_division = first_str
            continue
        if not has_title:
            continue

        record = {}
        for field in OUTPUT_SCHEMA:
            if field == "No.":
                record[field] = pd.NA
            elif field == "Division":
                record[field] = current_division
            elif field in STATUS_LABELS:
                record[field] = pd.NA
            elif field in ("Lat", "Long"):
                continue
            elif field == "Source Sheet":
                record[field] = sheet_name
            else:
                original_col_name = mapping.get(field)
                record[field] = row.get(original_col_name) if original_col_name else pd.NA

        if "Coordinates" in mapping:
            record["_TEMP_COORDINATES_RAW"] = row.get(mapping["Coordinates"])
        if "Lat" in mapping:
            record["_TEMP_LAT_RAW"] = row.get(mapping["Lat"])
        if "Long" in mapping:
            record["_TEMP_LONG_RAW"] = row.get(mapping["Long"])

        processed_records.append(record)

    if not processed_records:
        return pd.DataFrame(columns=OUTPUT_SCHEMA)

    unified = pd.DataFrame(processed_records)
    for col in OUTPUT_SCHEMA:
        if col not in unified.columns:
            unified[col] = pd.NA

    unified = populate_coordinates_from_temp(unified)
    unified = _drop_temp_coord_cols(unified)
    return unified[OUTPUT_SCHEMA]


def parse_flat(df, mapping, sheet_name):
    """Fallback: already one clean row per project, no blocks and no
    section dividers to collapse."""
    title_col = mapping.get("Project Title")
    if not title_col or title_col not in df.columns:
        return pd.DataFrame(columns=OUTPUT_SCHEMA)

    df_filtered = df[df[title_col].notna() & (df[title_col].astype(str).str.strip() != "")].copy()
    if df_filtered.empty:
        return pd.DataFrame(columns=OUTPUT_SCHEMA)

    unified_data = {}
    for field in OUTPUT_SCHEMA:
        if field == "No.":
            unified_data[field] = pd.NA
        elif field == "Division":
            unified_data[field] = df_filtered[mapping["Division"]] if mapping.get("Division") else sheet_name
        elif field in STATUS_LABELS:
            unified_data[field] = pd.NA
        elif field in ("Lat", "Long"):
            continue
        elif field == "Source Sheet":
            unified_data[field] = sheet_name
        else:
            original_col_name = mapping.get(field)
            unified_data[field] = df_filtered[original_col_name] if original_col_name and original_col_name in df_filtered.columns else pd.NA

    unified = pd.DataFrame(unified_data, index=df_filtered.index)

    if "Coordinates" in mapping:
        unified["_TEMP_COORDINATES_RAW"] = df_filtered.get(mapping["Coordinates"])
    if "Lat" in mapping:
        unified["_TEMP_LAT_RAW"] = df_filtered.get(mapping["Lat"])
    if "Long" in mapping:
        unified["_TEMP_LONG_RAW"] = df_filtered.get(mapping["Long"])

    unified = populate_coordinates_from_temp(unified)
    unified = _drop_temp_coord_cols(unified)
    return unified[OUTPUT_SCHEMA]


def normalize_division(name):
    if not isinstance(name, str):
        return name
    return DIVISION_NORMALIZATION.get(name.strip().upper(), name.strip())


def _assign_map_status(row):
    """Collapse the one-hot status columns into a single display label.
    Data sources that don't carry status at all (e.g. the cost-list
    format) leave all three as NA, which correctly falls through here."""
    for label in STATUS_LABELS:
        value = row.get(label)
        if pd.notna(value) and value == 1:
            return label
    return "Unknown"


def _compute_effective_funding(row):
    """Prefer the Revised amount when present and non-zero; otherwise fall
    back to the Original amount."""
    for col in (AMOUNT_REVISED_COL, AMOUNT_ORIGINAL_COL):
        raw = row.get(col)
        if pd.isna(raw):
            continue
        try:
            val = float(str(raw).replace(",", "").strip())
        except (ValueError, TypeError):
            continue
        if val != 0:
            return val
    return pd.NA


def _is_visible_sheet(ws):
    """Excel sheet_state is 'visible', 'hidden', or 'veryHidden'. Only
    sheets someone can actually see in Excel should be treated as
    candidate data -- a hidden sheet is usually a working copy, an old
    snapshot, or a scratch calculation, not the current source of truth."""
    return ws.sheet_state == "visible"


def process_workbook(uploaded_file, sheet_filter=None):
    """Scan every visible sheet (or only those in sheet_filter), auto-detect
    each sheet's layout, and normalize everything to OUTPUT_SCHEMA. Hidden
    and very-hidden sheets are always skipped.

    sheet_filter: optional iterable of sheet names to restrict scanning to
    (used by Legacy Tracker mode). None scans every visible sheet in the
    workbook.
    """
    file_bytes = uploaded_file.getvalue() if hasattr(uploaded_file, "getvalue") else uploaded_file.read()
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    all_results = []

    for sheet_name in wb.sheetnames:
        if sheet_filter is not None and sheet_name not in sheet_filter:
            continue

        ws = wb[sheet_name]
        if not _is_visible_sheet(ws):
            continue

        header_row = find_header_row(ws)
        if header_row is None:
            continue

        df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name, header=header_row - 1, dtype=str)
        df.columns = [str(c).strip() for c in df.columns]
        mapping = map_headers(list(df.columns))

        if "Project Title" not in mapping:
            continue

        structure = detect_structure(df, mapping)

        if structure == "merged_status":
            unified = parse_merged_status(df, mapping, sheet_name)
        elif structure == "section_headers":
            unified = parse_section_headers(df, mapping, sheet_name)
        else:
            unified = parse_flat(df, mapping, sheet_name)

        if unified.empty:
            continue

        unified["Division"] = unified["Division"].apply(normalize_division)
        all_results.append(unified)

    if not all_results:
        return pd.DataFrame(columns=OUTPUT_SCHEMA)

    combined = pd.concat(all_results, ignore_index=True)
    combined["No."] = combined.groupby("Division").cumcount() + 1
    return combined[OUTPUT_SCHEMA]


@st.cache_data
<<<<<<< HEAD
def list_candidate_sheets(uploaded_file):
    """Return every VISIBLE sheet name in the workbook that has a
    recognizable project-list header row. Used to build the snapshot/sheet
    picker for Cost List / Any Template mode, since a workbook can contain
    several dated snapshots of the same projects rather than one sheet per
    division -- and often also hidden working copies that shouldn't be
    offered as choices at all."""
    file_bytes = uploaded_file.getvalue() if hasattr(uploaded_file, "getvalue") else uploaded_file.read()
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    candidates = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if _is_visible_sheet(ws) and find_header_row(ws) is not None:
            candidates.append(sheet_name)
    return candidates


@st.cache_data
def load_and_clean_data(uploaded_file, ingestion_mode, selected_sheet=None):
    """Ingest the workbook under the chosen mode and finish deriving the
    fields the rest of the app relies on (numeric Lat/Long, a single
    display status, a single effective funding figure).

    Rows without parseable coordinates are kept here (not dropped) --
    they're the candidates for the "Resolve Missing Coordinates" geocoding
    step in the sidebar. The app still only maps/tracks KPIs for rows that
    end up with coordinates; that filtering happens later, after geocoding
    has had a chance to fill gaps in. Rows without status simply display
    as 'Unknown'.

=======
def load_and_clean_data(uploaded_file, ingestion_mode, selected_sheet=None):
    """Ingest the workbook under the chosen mode and finish deriving the
    fields the rest of the app relies on (numeric Lat/Long, a single
    display status, a single effective funding figure).

    Rows without parseable coordinates are kept here (not dropped) --
    they're the candidates for the "Resolve Missing Coordinates" geocoding
    step in the sidebar. The app still only maps/tracks KPIs for rows that
    end up with coordinates; that filtering happens later, after geocoding
    has had a chance to fill gaps in. Rows without status simply display
    as 'Unknown'.

>>>>>>> 377714d600e0645afde87abb5929f0b11189bc5a
    In Legacy Tracker mode, every CEST/LGIA/SSCP sheet is combined (they
    are genuinely different divisions). In Cost List / Any Template mode,
    only `selected_sheet` is loaded -- a workbook in this format is often
    several dated snapshots of the *same* projects, not separate
    divisions, so combining all sheets would multiply-count every project.
    """
    if ingestion_mode == "legacy":
        sheet_filter = DATA_SHEETS
    elif selected_sheet is not None:
        sheet_filter = [selected_sheet]
    else:
        sheet_filter = None

    combined = process_workbook(uploaded_file, sheet_filter=sheet_filter)

    if combined.empty:
        return combined

    combined["Lat"] = pd.to_numeric(combined["Lat"], errors="coerce")
    combined["Long"] = pd.to_numeric(combined["Long"], errors="coerce")
    combined["Map_Status"] = combined.apply(_assign_map_status, axis=1)
    combined[EFFECTIVE_FUNDING_COL] = combined.apply(_compute_effective_funding, axis=1)

    return combined


# ==========================================
# 2.5 GEOCODING ENGINE (MAPBOX)
# ==========================================
# Task 7.1: Provider Setup & Security
# ------------------------------------
# Requires a Mapbox access token stored in .streamlit/secrets.toml:
#   MAPBOX_API_KEY = "pk.xxxxxxxx"
# Never hardcode the token in source. If it's missing, geocoding is
# disabled in the UI (see Task 7.4) rather than the app crashing.
MAPBOX_GEOCODING_URL = "https://api.mapbox.com/geocoding/v5/mapbox.places/{query}.json"
GEOCODE_REGION_CONTEXT = "Davao Region, Philippines"
GEOCODE_COUNTRY_CODE = "ph"
# (min_lon, min_lat, max_lon, max_lat) -- covers Davao del Norte, Davao del
# Sur, Davao Oriental, Davao de Oro, and Davao Occidental with a small
# buffer. Used both as a hard Mapbox `bbox` filter and as a client-side
# sanity check on whatever comes back, so a fuzzy agency-name match in
# Luzon or Visayas gets rejected instead of silently plotted.
DAVAO_REGION_BBOX = (125.0, 5.5, 126.7, 8.1)  # min_lon, min_lat, max_lon, max_lat


def get_mapbox_token():
    """Read the Mapbox token from Streamlit secrets. Returns None (rather
    than raising) if it isn't configured, so the caller can degrade
    gracefully instead of crashing the whole app."""
    try:
        return st.secrets.get("MAPBOX_API_KEY")
    except Exception:
        return None


# Task 7.2: Address Contextualization Logic
# ------------------------------------------
def build_geocode_query(row):
    """Build a search string for the geocoder out of whatever location
    signal a row actually has. Prefers 'Implementing Agency' (which also
    covers 'Implementing Partner' -- both map to this canonical field) since
    that's a searchable organization/site name, and falls back to 'Location'
    only if no agency/partner is recorded. Appends the region context so a
    bare name like 'DOST XI Regional Office' resolves to somewhere in Davao
    Region instead of matching globally."""
    agency = clean_missing(row.get("Implementing Agency", ""), fallback="")
    location = clean_missing(row.get("Location", ""), fallback="")

    base = None
    if agency:
        base = agency
    elif location:
        base = location

    if not base:
        return None

    base = base.strip()
    if "davao" not in base.lower() and "philippines" not in base.lower():
        base = f"{base}, {GEOCODE_REGION_CONTEXT}"

    return base


# Task 7.3: Cached Fetching & Rate Limit Handling
# -------------------------------------------------
@st.cache_data(show_spinner=False)
def geocode_address_mapbox(query, api_token):
    """Resolve one query string to (lat, long) via the Mapbox Geocoding
    API. Cached by (query, token) so a query already resolved this
    session never re-hits the API -- important both for cost and for
    respecting rate limits on repeated runs over the same dataset.

    Results are hard-constrained to DAVAO_REGION_BBOX (both via Mapbox's
    own `bbox` parameter and a client-side re-check) so a fuzzy match on
    an agency/office name common across the Philippines -- e.g. multiple
    unrelated offices sharing a generic name -- can't resolve to a
    same-named place in Luzon or Visayas. Returns None on any failure (no
    match, out-of-region match, network error, bad response) so the
    caller can add it to the manual-review fallback list."""
    if not query or not api_token:
        return None

    min_lon, min_lat, max_lon, max_lat = DAVAO_REGION_BBOX
    url = MAPBOX_GEOCODING_URL.format(query=requests.utils.quote(query))
    params = {
        "access_token": api_token,
        "limit": 1,
        "country": GEOCODE_COUNTRY_CODE,
        "bbox": f"{min_lon},{min_lat},{max_lon},{max_lat}",
        "proximity": f"{DAVAO_CENTER[1]},{DAVAO_CENTER[0]}",  # Mapbox wants lon,lat
    }

    try:
        response = requests.get(url, params=params, timeout=10)
        response.raise_for_status()
        data = response.json()
    except (requests.RequestException, ValueError):
        return None
    finally:
        # Gentle throttle -- well within Mapbox's free-tier rate limits,
        # just avoids hammering the API on large batches.
        time.sleep(0.1)

    features = data.get("features") or []
    if not features:
        return None

    # Mapbox returns [longitude, latitude].
    lon, lat = features[0]["center"]

    # Belt-and-suspenders: reject anything outside the region even if the
    # bbox param somehow let it through (e.g. an older Mapbox dataset
    # ignoring bbox for an exact-name hit).
    if not (min_lon <= lon <= max_lon and min_lat <= lat <= max_lat):
        return None

    return (lat, lon)


def resolve_missing_coordinates(df, api_token):
    """Task 7.4 core loop (UI lives in the sidebar, see MAIN APPLICATION
    LOGIC): geocode every row missing Lat/Long in place. Returns the
    updated df plus a list of rows that couldn't be resolved, so DOST
    staff can review them manually rather than the app silently dropping
    them."""
    df = df.copy()
    missing_mask = df["Lat"].isna() | df["Long"].isna()
    missing_idx = df[missing_mask].index.tolist()

    failures = []
    progress = st.sidebar.progress(0, text="Resolving coordinates...")

    for i, idx in enumerate(missing_idx):
        row = df.loc[idx]
        query = build_geocode_query(row)
        coords = geocode_address_mapbox(query, api_token) if query else None

        if coords:
            df.at[idx, "Lat"] = coords[0]
            df.at[idx, "Long"] = coords[1]
        else:
            failures.append({
                "Project Title": row.get("Project Title", "Untitled"),
                "Division": row.get("Division", ""),
                "Query Used": query or "(no Location or Implementing Agency to search)",
            })

        progress.progress((i + 1) / len(missing_idx), text=f"Resolving {i + 1}/{len(missing_idx)}...")

    progress.empty()
    return df, failures


# ==========================================
# 3. MAP GENERATION
# ==========================================
def _build_standard_pin_html(hex_color, scale=1.0):
    """Build the HTML for a standard map pin with a specific color."""
    return f"""
    <div style="display: flex; justify-content: center; align-items: center; width: 40px; height: 40px; transform: scale({scale}); transform-origin: bottom center;">
        <svg viewBox="0 0 24 24" fill="{hex_color}" xmlns="http://www.w3.org/2000/svg" style="width: 30px; height: 30px; filter: drop-shadow(2px 4px 6px rgba(0,0,0,0));">
            <path d="M12 2C8.13 2 5 5.13 5 9c0 5.25 7 13 7 13s7-7.75 7-13c0-3.87-3.13-7-7-7zm0 9.5c-1.38 0-2.5-1.12-2.5-2.5s1.12-2.5 2.5-2.5 2.5 1.12 2.5 2.5-1.12 2.5-2.5 2.5z"/>
        </svg>
    </div>
    """


def _build_badge_html(abbrev, hex_color, scale=1.0):
    """Build the HTML for a marker's abbreviation badge."""
    return f"""
    <div style="
        display: inline-flex;
        align-items: center;
        width: max-content;
        background-color: white;
        border: 2px solid {hex_color};
        border-radius: 12px;
        padding: 4px 8px;
        font-size: 11px;
        font-family: Arial, sans-serif;
        font-weight: bold;
        color: #2c3e50;
        box-shadow: 0px 3px 6px rgba(0,0,0,0.3);
        white-space: nowrap;
        cursor: pointer;
        transform: scale({scale});
        transform-origin: center center;
    ">
        <div style="width: 8px; height: 8px; border-radius: 50%; background-color: {hex_color}; margin-right: 5px;"></div>
        {abbrev}
    </div>
    """


_ABBREV_STOPWORDS = {"of", "the", "and", "for", "in", "to", "a", "an", "on", "at", "&"}


def derive_project_abbreviation(title, max_letters=6):
    """Build a short badge label out of a project's title when no
    'Name Abbreviation' was supplied by the source file -- initials of the
    significant words, e.g. 'Community-Based Skills Training Program' ->
    'CBSTP'. Falls back to the first few letters of the title itself if it
    can't extract multiple words, and to 'N/A' if there's no title at all."""
    if not isinstance(title, str) or not title.strip():
        return "N/A"

    words = re.findall(r"[A-Za-z0-9]+", title)
    significant = [w for w in words if w.lower() not in _ABBREV_STOPWORDS] or words
    if not significant:
        return "N/A"

    if len(significant) == 1:
        return significant[0][:max_letters].upper()

    return "".join(w[0].upper() for w in significant[:max_letters])


def get_marker_label(row):
    """Return the badge label to render on the map: the source file's own
    'Name Abbreviation' when it actually has one, otherwise an abbreviation
    derived from the Project Title. This is what fixes newly-geocoded
    projects (and any Cost List / Any Template row) showing up as a blank
    or 'nan' badge on the map."""
    raw = row.get("Name Abbreviation")
    if pd.notna(raw) and str(raw).strip():
        return str(raw).strip()
    return derive_project_abbreviation(row.get("Project Title"))


def add_legend(folium_map):
    legend_html = '''
    <div style="
        position: absolute; 
        bottom: 30px; 
        right: 10px; 
        z-index: 9999; 
        background-color: white; 
        padding: 15px; 
        border-radius: 8px; 
        border: 2px solid rgba(0,0,0,0.1);
        box-shadow: 0px 3px 6px rgba(0,0,0,0.3);
        font-family: 'Montserrat', Arial, sans-serif;
        font-size: 12px;
        ">
        <h4 style="margin: 0 0 10px 0; font-size: 14px; font-weight: bold; color: #2c3e50;">Project Divisions</h4>
    '''
    for div, color in DIVISION_COLORS.items():
        legend_html += f'''
        <div style="display: flex; align-items: center; margin-bottom: 8px;">
            <div style="width: 14px; height: 14px; background-color: {color}; border-radius: 50%; margin-right: 8px;"></div>
            <span style="color: #2c3e50; font-weight: 600;">{div}</span>
        </div>
        '''
    legend_html += f'''
        <div style="display: flex; align-items: center;">
            <div style="width: 14px; height: 14px; background-color: {DEFAULT_MARKER_COLOR}; border-radius: 50%; margin-right: 8px;"></div>
            <span style="color: #2c3e50; font-weight: 600;">Other</span>
        </div>
    </div>
    '''
    folium_map.get_root().html.add_child(folium.Element(legend_html))


def create_map(df, pin_style="Abbreviation Badge", enable_clustering=True, show_legend=True, scale_by_funding=False):
    """Render the Davao region map with a clustered marker per project.
    Rows without coordinates (e.g. from a source file that has no
    location data) are simply skipped -- they still show up in the table,
    KPIs, and exports, just not on the map."""
    davao_map = folium.Map(location=DAVAO_CENTER, zoom_start=DAVAO_ZOOM_START)
    if enable_clustering:
        marker_parent = MarkerCluster().add_to(davao_map)
    else:
        marker_parent = davao_map
        
    mappable_df = df.dropna(subset=["Lat", "Long"])

    max_funding = 0
    if scale_by_funding and EFFECTIVE_FUNDING_COL in df.columns:
        funding_numeric = pd.to_numeric(df[EFFECTIVE_FUNDING_COL], errors='coerce')
        if not funding_numeric.empty:
            max_funding = funding_numeric.max()

    for _, row in mappable_df.iterrows():
        abbrev = get_marker_label(row)
        division = str(row.get("Division", "N/A"))
        hex_color = DIVISION_COLORS.get(division, DEFAULT_MARKER_COLOR)

        scale = 1.0
        if scale_by_funding and max_funding > 0:
            funding = pd.to_numeric(row.get(EFFECTIVE_FUNDING_COL), errors='coerce')
            if pd.notna(funding) and funding > 0:
                # Scale between 0.2x and 1.2x based on area (sqrt of funding)
                scale = 0.2 + (1.2 - 0.2) * ((funding ** 0.5) / (max_funding ** 0.5))

        if pin_style == "Standard Color Pin":
            icon = DivIcon(
                icon_anchor=(15, 30),
                html=_build_standard_pin_html(hex_color, scale),
                class_name="custom-pin",
            )
        else:
            icon = DivIcon(
                icon_anchor=(0, 0),
                html=_build_badge_html(abbrev, hex_color, scale),
                # Removes Leaflet's default icon size/overflow constraints
                # so the badge can render at its natural size.
                class_name="custom-badge",
            )

        folium.Marker(
            location=[row["Lat"], row["Long"]],
            tooltip="📍 View project details",
            icon=icon,
        ).add_to(marker_parent)

    if show_legend:
        add_legend(davao_map)

    return davao_map


# ==========================================
# 4. PAGE CONFIGURATION
# ==========================================
st.set_page_config(
    page_title="DOST-Davao Project Mapper",
    page_icon="assets/dost_icon.png",
    layout="wide",
)

st.markdown("""
    <style>
 
        /* ==========================================
           1. FONT IMPORT
           ========================================== */
        /* Montserrat weights: 400=Regular, 700=Bold, 900=Black */
        @import url('https://fonts.googleapis.com/css2?family=Montserrat:wght@400;700;900&display=swap');
 
 
        /* ==========================================
           2. TYPOGRAPHY
           ========================================== */
        /* Scale the whole app down ~5%. Streamlit's spacing, padding, and
           font sizes are almost entirely rem-based, so shrinking the root
           font-size scales everything proportionally. This is used instead
           of transform: scale(), which clips iframes (the Folium map) and
           throws off click coordinates. */
        html {
            font-size: 90% !important;
        }

        /* Base font for the whole app */
        html, body, [class*="css"] {
            font-family: 'Montserrat', 'Arial', sans-serif !important;
            font-weight: 400 !important;
        }
 
        /* Sub-headers use Bold */
        h2, h3, h4, h5, h6 {
            font-family: 'Montserrat', 'Arial', sans-serif !important;
            font-weight: 700 !important;
        }
 
        /* Main H1 header uses Black (900 is the max valid CSS weight) */
        h1 {
            font-family: 'Montserrat', 'Arial', sans-serif !important;
            font-weight: 900 !important;
        }
 
 
        /* ==========================================
           3. LAYOUT SPACING
           ========================================== */
        /* Trim default top/bottom padding on the main content area */
        .block-container {
            padding-top: 1.2rem;
            padding-bottom: 0rem;
        }
 
        /* Remove top padding above sidebar content */
        [data-testid="stSidebar"] [data-testid="stSidebarUserContent"] {
            padding-top: 0rem;
        }
 
        /* Collapse the invisible sidebar header spacing */
        [data-testid="stSidebarHeader"] {
            padding-top: 0rem !important;
            padding-bottom: 0rem !important;
            min-height: 0px !important;
        }
 
        /* Belt-and-suspenders: force sidebar content padding to zero */
        [data-testid="stSidebarUserContent"] {
            padding-top: 0rem !important;
        }
 
        /* Pull sidebar image up to eat leftover space above it */
        [data-testid="stSidebar"] .stImage {
            margin-top: -2.5rem;
        }

        /* Divider */
        .st-emotion-cache-17ta2sm hr {
            padding: 0 !important;
            color: inherit !important;
            border-width: medium medium 1px !important;
            border-style: none none solid !important;
            border-color: currentcolor currentcolor #FFFFFF !important;  
            border-image: none !important;
        }
 
        /* ==========================================
           4. FILE UPLOADER
           ========================================== */
        /* Force all text inside the uploader to white */
        [data-testid="stFileUploader"] div,
        [data-testid="stFileUploader"] span,
        [data-testid="stFileUploader"] small,
        [data-testid="stFileUploader"] label,
        [data-testid="stFileUploader"] p {
            color: #FFFFFF !important;
        }
 
        /* Dropzone as a column. align-items: stretch lets the uploaded-file
           row go full width (fixes filename truncation) — the instructions
           text and upload button opt out of stretching via align-self. */
        [data-testid="stFileUploaderDropzone"] {
            width: 100% !important;
            display: flex !important;
            flex-direction: column !important;
            align-items: stretch !important;
            justify-content: center !important;
        }
 
        /* Center the empty-state instructions text */
        [data-testid="stFileUploaderDropzoneInstructions"] {
            align-self: center !important;
            text-align: center !important;
            margin-bottom: -3.5rem !important;
        }
 
        /* Center the Upload button */
        [data-testid="stFileUploaderDropzone"] [data-testid="stBaseButton-secondary"] {
            width: 100% !important;
            align-self: center !important;
        }
 
        /* "Add files" button, present only when accept_multiple_files=True */
        [data-testid="stBaseButton-borderlessIcon"] {
            color: #FFFFFF !important;
            margin-top: -1.7rem !important;
        }
 
        [data-testid="stBaseButton-borderlessIcon"] [data-testid="stIconMaterial"] {
            color: transparent !important;
        }
 
        .st-emotion-cache-oeolxv { /* for svg icon in file uploader */
            background-color: #27343a !important;
            outline: 2px solid #1f292e !important;
        }

        .st-emotion-cache-pedct3 button { /* for 'x' icon in file uploader */   
            color: #27343a !important;
            background-color: #FFFFFF !important;
        }
            
        /* Let the uploaded filename take its natural width instead of truncating */
        [data-testid="stFileUploaderDropzone"] [data-testid="stFileUploaderFileName"] {
            max-width: none !important;
            overflow: visible !important;
            text-overflow: unset !important;
            white-space: normal !important;
            word-break: break-word !important;
        }
 
 
        /* ==========================================
           5. SIDEBAR TEXT COLOR
           ========================================== */
        /* Force all sidebar text elements to white */
        [data-testid="stSidebar"] h1,
        [data-testid="stSidebar"] h2,
        [data-testid="stSidebar"] h3,
        [data-testid="stSidebar"] p,
        [data-testid="stSidebar"] label,
        [data-testid="stSidebar"] span {
            color: #FFFFFF !important;
        }
 
        /* Metric values need a separate override */
        [data-testid="stSidebar"] [data-testid="stMetricValue"] {
            color: #FFFFFF !important;
        }
 
 
        /* ==========================================
           6. SIDEBAR ELEMENT SPACING
           ========================================== */
 
        /* Give the sidebar itself some horizontal padding so content
           doesn't sit flush against the edges */
        [data-testid="stSidebarUserContent"] {
            padding-left: 0.5rem !important;
            padding-right: 0.5rem !important;
        }
 
        /* ==========================================
           7. CHROME / UI CLEANUP
           ========================================== */
        header {background-color: transparent !important;}
        #MainMenu {visibility: hidden;}
        .stDeployButton {display: none;}

        .st-e3 {
            background-color: transparent !important;
        }
 
 
        # /* ==========================================
        #    8. LOCK SIDEBAR SCROLLING
        #    ========================================== */
        # /* Force the sidebar container to hide ALL overflowing content */
        # [data-testid="stSidebar"] {
        #     overflow-x: hidden !important;
        #     overflow-y: hidden !important;
        # }
 
        # /* Streamlit sometimes wraps sidebar content in an inner scrolling div */
        # [data-testid="stSidebar"] > div:first-child {
        #     overflow-x: hidden !important;
        #     overflow-y: hidden !important;
        # }
 
        # /* Hide the scrollbar visual across all browsers */
        # [data-testid="stSidebar"] ::-webkit-scrollbar {
        #     display: none !important;
        #     width: 0px !important;
        #     height: 0px !important;
        # }
 
        # [data-testid="stSidebar"] * {
        #     scrollbar-width: none !important; /* Firefox */
        #     -ms-overflow-style: none !important; /* IE and Edge */
        # }
 
 
        /* ==========================================
           9. MAP UI ENHANCEMENTS
           ========================================== */
        /* Target the Folium iframe to round corners and add a soft shadow */
        iframe {
            border-radius: 16px !important;
            box-shadow: -5px 6px 26px 0px rgba(0,0,0,0.1) !important;
            overflow: hidden !important;
        
        }
            
        /* ==========================================
           10. CHAT UI FIXES
           ========================================== */
        /* 1. Force the main chat input container to a crisp white with a DOST blue outline */
        [data-testid="stChatInput"] > div {
            color: #2C3E50 !important; /* Sets base text and caret color to dark gray */
            background-color: #FFFFFF !important;
            outline: 2px solid #00AEEF !important;
        }
            
        /* 2. Force the actual typed text to be readable */
        [data-testid="stChatInput"] textarea {
            color: #000000  !important; /* Dark gray typing text */
            background-color: #FFFFFF !important; /* Lets the white container show through safely */
        }

        /* 3. Style the 'Send' paper airplane icon to match your DOST Blue outline */
        [data-testid="stChatInput"] button svg {
            fill: #FFFFFF !important; /* Blue icon instead of invisible white */
        }
            
        /* 4. Completely un-anchor the chat input wrapper from the bottom */
        [data-testid="stBottomBlock"], 
        [data-testid="stBottom"],
        .stChatFloatingInputContainer,
        [data-testid="stChatInput"] {
            position: static !important;
            bottom: auto !important;
            background-color: transparent !important;
        }


        /* ==========================================
           11. CHAT AVATAR COLORS
           ========================================== */
        /* User messages: dark slate */
        [data-testid="stChatMessageAvatarUser"] {
            background-color: #222D32 !important;
        }

        /* Assistant messages: DOST blue */
        [data-testid="stChatMessageAvatarAssistant"] {
            background-color: #00AEEF !important;
        }

        /* ==========================================
           12. SIDEBAR ELEMENTS COLOR FIX (multiselect / selectbox)
           ========================================== */
        /* The previous version of this fix targeted Streamlit's
           auto-generated .st-emotion-cache-* / .st-XX class names, captured
           from a browser inspector at one point in time. Those hashes come
           from Emotion (CSS-in-JS) and are NOT stable across Streamlit
           versions or even separate rebuilds -- which is exactly why the
           filter widgets intermittently reverted to an unstyled white
           background that blended into the page. BaseWeb (the underlying
           component library Streamlit's select/multiselect widgets are
           built on) instead exposes `data-baseweb` attributes that don't
           change with the build hash, so we target those.
        */

        /* The select/multiselect input box itself: keep it a crisp white
           box so it reads clearly against the dark sidebar, regardless of
           which internal class names this Streamlit build happens to use. */
        [data-testid="stSidebar"] [data-baseweb="select"] > div {
            background-color: #FFFFFF !important;
            border-color: #FFFFFF !important;
        }
        [data-testid="stSidebar"] [data-baseweb="select"] input {
            color: #2C3E50 !important;
        }

        /* Selected-value "pills" inside multiselect widgets (e.g. the
           Division / Status filters) */
        [data-testid="stSidebar"] [data-baseweb="tag"] {
            background-color: #00AEEF !important;
            color: #FFFFFF !important;
        }
        [data-testid="stSidebar"] [data-baseweb="tag"] span {
            color: #FFFFFF !important;
        }
        [data-testid="stSidebar"] [data-baseweb="tag"] svg {
            fill: #FFFFFF !important;
        }

        /* The dropdown options list renders in a portal at the document
           root, not nested inside the sidebar, so it needs its own
           (unscoped) selector rather than [data-testid="stSidebar"] ... */
        [data-baseweb="popover"] [data-baseweb="menu"] {
            background-color: #FFFFFF !important;
        }
        [data-baseweb="popover"] [data-baseweb="menu"] li,
        [data-baseweb="popover"] [data-baseweb="menu"] li * {
            color: #2C3E50 !important;
        }

        [role=radiogroup] {
            margin-left: 1rem !important;
        }

        [data-testid="stTooltipIcon"] svg {
            stroke: #FFFFFF !important;
            stroke-width: 2.25 !important;
        }

        .st-by {
            background-color: #FFFFFF !important;
        }

        .st-g6 {
            background-color: #FFFFFF !important;
        }

        /* ==========================================
           13. SLIDER TWEAKS
           ========================================== */
        /* Hide the min/max bounds that fade in and out on the slider */
        [data-testid="stTickBar"] {
            display: none !important;
        }

        /* ==========================================
           14. FULLSCREEN MAP DIALOG OVERRIDE
           ========================================== */
        /* Force the Fullscreen Map View dialog to exceed Streamlit's "large" max-width */
        div[data-testid="stDialog"] [role="dialog"] {
            width: 95vw !important;
            max-width: 95vw !important;
            margin-top: -2.5vh !important; /* Pull the dialog up higher */
            align-self: flex-start !important; /* Anchor it to the top if it's a flex container */
        }

    </style>
""", unsafe_allow_html=True)
 
st.markdown("### DOST-Davao Project Mapper 📍")

# ==========================================
# 5. HELPERS FOR DISPLAY FORMATTING
# ==========================================
def clean_missing(value, fallback="TBA / Not specified"):
    """Replace pandas' NaN/NaT text representations with a readable fallback."""
    text = str(value)
    return fallback if text.lower() in MISSING_VALUE_TOKENS else text


def format_currency(amount):
    """Format a peso amount, abbreviating to millions when large."""
    if amount >= 1_000_000:
        return f"₱{amount / 1_000_000:.2f}M"
    return f"₱{amount:,.0f}"


# ==========================================
# 6. MODAL DIALOGS
# ==========================================
def render_project_details_content(clicked_projects):
    st.success(f"Found {len(clicked_projects)} project(s) at this location.")
    st.divider()

    for _, row in clicked_projects.iterrows():
        title = str(row.get(GROUP_KEY_COL, "Unnamed Project"))
        division = str(row.get("Division", "N/A"))
        status = str(row.get("Map_Status", "Unknown"))
        
        funding_raw = row.get(EFFECTIVE_FUNDING_COL, "Not specified")
        try:
            funding_val = float(str(funding_raw).replace(',', ''))
            display_funding = f"₱{funding_val:,.2f}" if funding_val % 1 != 0 else f"₱{int(funding_val):,}"
        except (ValueError, TypeError):
            display_funding = f"₱{funding_raw}" if funding_raw != "Not specified" else "Not specified"

        agency = row.get("Implementing Agency", "Not specified")
        proponent = str(row.get("Name of Proponent", "Not specified"))
        coordinator = clean_missing(row.get("DOST Coordinator", "Not specified"))
        location = clean_missing(row.get("Location", "Not specified"))
        beneficiaries = clean_missing(row.get("Beneficiaries", "Not specified"))
        remarks = str(row.get("Remarks", "No remarks provided."))

        date_approved = clean_missing(row.get(DATE_APPROVED_COL, "Not specified"))
        date_end = clean_missing(row.get(DATE_END_COL, "Not specified"))

        st.header(title)
        st.subheader(f"**Division:** {division} | **Status:** {status}")

        col1, col2 = st.columns(2)
        with col1:
            st.markdown(f"**🏢 Agency:** {agency}")
            st.markdown(f"**👤 Proponent:** {proponent}")
            st.markdown(f"**💰 Funding:** {display_funding}")
            if coordinator != "Not specified":
                st.markdown(f"**🧑‍💼 DOST Coordinator:** {coordinator}")
        with col2:
            # [:10] strips any trailing " 00:00:00" timestamp from date strings.
            st.markdown(f"**📅 Date Approved:** {date_approved[:10]}")
            st.markdown(f"**🏁 End/Target Date:** {date_end[:10]}")
            if location != "Not specified":
                st.markdown(f"**📍 Location:** {location}")
            if beneficiaries != "Not specified":
                st.markdown(f"**🎯 Beneficiaries:** {beneficiaries}")

        st.info(f"**Remarks:** {remarks}")
        st.divider()


@st.dialog("📋 Located Project Details", width="large")
def show_project_details(clicked_projects):
    """Display full details for every project at a clicked map location."""
    render_project_details_content(clicked_projects)


@st.dialog("📊 Raw Data Explorer", width="large")
def show_raw_data(df):
    """Display the filtered dataset in a plain, sortable table."""
    st.caption(f"Currently viewing {len(df)} filtered projects.")
    st.dataframe(df, width="stretch")


@st.dialog("🗺️ Fullscreen Map View", width="large")
def show_fullscreen_map(df, pin_style, enable_clustering, show_legend, scale_by_funding):
    """Render the map in a large modal window for better visibility."""
    project_map = create_map(df, pin_style, enable_clustering, show_legend, scale_by_funding)
    map_key = f"fullscreen_map_{show_legend}_{enable_clustering}_{pin_style}_{scale_by_funding}"
    map_data = st_folium(
        project_map,
        use_container_width=True,
        height=570,
        returned_objects=["last_object_clicked"],
        key=map_key
    )
    
    clicked = map_data.get("last_object_clicked")
    if clicked:
        click_key = (round(clicked["lat"], 6), round(clicked["lng"], 6))
        
        # Only show the spinner transition if the user clicked a NEW marker
        if st.session_state.get("fullscreen_last_click") != click_key:
            st.session_state["fullscreen_last_click"] = click_key
            with st.spinner("Fetching project details..."):
                time.sleep(3.5)
                
        clicked_projects = df[
            (abs(df["Lat"] - clicked["lat"]) < CLICK_TOLERANCE_DEGREES)
            & (abs(df["Long"] - clicked["lng"]) < CLICK_TOLERANCE_DEGREES)
        ]
        if not clicked_projects.empty:
            st.markdown("### 📋 Clicked Location Details")
            render_project_details_content(clicked_projects)


@st.dialog("🖼️ Map Export Options", width="large")
def show_export_dialog(df, pin_style, enable_clustering, show_legend, scale_by_funding):
    """Configure, preview, and download the high-res map export."""
    if not EXPORT_AVAILABLE:
        st.warning("⚠️ High-Res Map Export is currently disabled. Please install `selenium`, `webdriver-manager`, and `pillow` to enable this feature.")
        return

    st.write("Configure your map export settings below:")
    
    col1, col2 = st.columns(2)
    with col1:
        zoom_level = st.slider("Zoom Level", min_value=8, max_value=18, value=13)
    with col2:
        title = st.text_input("Report Title", value="DOST-Davao Regional Project Mapping Report")
        
    if st.button("Generate Preview", width="stretch"):
        with st.spinner("Capturing map and generating report (this may take a few seconds)..."):
            # Create a dedicated map for export with the selected zoom and use the fixed DAVAO_CENTER
            export_map = folium.Map(location=DAVAO_CENTER, zoom_start=zoom_level, zoom_control=False)
            if enable_clustering:
                marker_parent = MarkerCluster().add_to(export_map)
            else:
                marker_parent = export_map

            max_funding = 0
            if scale_by_funding and EFFECTIVE_FUNDING_COL in df.columns:
                funding_numeric = pd.to_numeric(df[EFFECTIVE_FUNDING_COL], errors='coerce')
                if not funding_numeric.empty:
                    max_funding = funding_numeric.max()

            for _, row in df.dropna(subset=["Lat", "Long"]).iterrows():
                abbrev = get_marker_label(row)
                division = str(row.get("Division", "N/A"))
                hex_color = DIVISION_COLORS.get(division, DEFAULT_MARKER_COLOR)

                scale = 1.0
                if scale_by_funding and max_funding > 0:
                    funding = pd.to_numeric(row.get(EFFECTIVE_FUNDING_COL), errors='coerce')
                    if pd.notna(funding) and funding > 0:
                        scale = 0.7 + (2.5 - 0.7) * ((funding ** 0.5) / (max_funding ** 0.5))

                if pin_style == "Standard Color Pin":
                    icon = DivIcon(
                        icon_anchor=(15, 30),
                        html=_build_standard_pin_html(hex_color, scale),
                        class_name="custom-pin",
                    )
                else:
                    icon = DivIcon(
                        icon_anchor=(0, 0),
                        html=_build_badge_html(abbrev, hex_color, scale),
                        class_name="custom-badge",
                    )

                folium.Marker(
                    location=[row["Lat"], row["Long"]],
                    icon=icon,
                ).add_to(marker_parent)
            
            if show_legend:
                add_legend(export_map)

            # Hide scrollbar in the export map
            export_map.get_root().header.add_child(folium.Element("<style>body, html { margin:0; padding:0; overflow: hidden !important; }</style>"))
            
            raw_image = capture_map_screenshot(export_map)
            
            active_divs = df["Division"].unique().tolist()
            active_stats = df["Map_Status"].unique().tolist()
            
            branded_bytes = create_branded_export(raw_image, active_divs, active_stats, title)
            
            # Store in session_state to persist between interactions in dialog
            st.session_state["export_preview_bytes"] = branded_bytes

    if "export_preview_bytes" in st.session_state:
        st.image(st.session_state["export_preview_bytes"], width="stretch")
        st.download_button(
            label="📥 Download High-Res Map (PNG)",
            data=st.session_state["export_preview_bytes"],
            file_name="DOST_Davao_Map_Export.png",
            mime="image/png",
            width="stretch"
        )


# ==========================================
# 7. SIDEBAR: BRANDING & FILE UPLOAD
# ==========================================
st.sidebar.image("assets/dost_davao_logo.png", width="stretch")

data_format_label = st.sidebar.radio(
    "Data Format:",
    options=list(INGESTION_MODES.keys()),
    index=0,
    help=(
        "Select the option that matches your Excel file.\n\n"
        "• Division-Based Template – For files organized into separate "
        "CEST, LGIA, and SSCP worksheets.\n\n"
        "• General Template – For project lists, project cost reports, "
        "and other supported DOST Excel templates. The application will "
        "automatically recognize the file format.\n\n"
        "Note: Only projects with valid coordinates can be displayed on the map."
    ),
)
ingestion_mode = INGESTION_MODES[data_format_label]

uploaded_file = st.sidebar.file_uploader(
    "Upload DOST-Davao Excel File (.xlsx)",
    type=["xlsx"],
)

selected_sheet = None
if uploaded_file is not None and ingestion_mode == "auto":
    candidate_sheets = list_candidate_sheets(uploaded_file)
    if not candidate_sheets:
        st.sidebar.error("⚠️ No sheet with a recognizable project list was found in this file.")
    elif len(candidate_sheets) == 1:
        selected_sheet = candidate_sheets[0]
    else:
        selected_sheet = st.sidebar.selectbox(
            "Select Worksheet:",
            options=candidate_sheets,
            index=len(candidate_sheets) - 1,
            help=(
                "This workbook contains multiple worksheets.\n\n"
                "Select the worksheet you want to display. "
                "Only the selected worksheet will be loaded and shown on the map."
            ),
        )


# ==========================================
# 8. SIDEBAR: FILTERS & KPI DASHBOARD
# ==========================================
def format_export_data(df):
    """Clean and reorder the dataframe for stakeholder presentation."""
    export_df = df.copy()
    
    # 1. Drop internal logic & redundant columns
    cols_to_drop = list(STATUS_LABELS) + ["Source Sheet"]
    export_df = export_df.drop(columns=[c for c in cols_to_drop if c in export_df.columns])
    
    # 2. Rename columns for professionalism
    rename_map = {
        "Map_Status": "Status",
        "Lat": "Latitude",
        "Long": "Longitude",
        EFFECTIVE_FUNDING_COL: "Funding (Effective)",
    }
    export_df = export_df.rename(columns=rename_map)
    
    # 3. Reorder: Move Status, Remarks, Lat, Long after the Date columns
    current_cols = list(export_df.columns)
    
    # Pull out our target columns so we can safely reposition them
    target_order = ["Status", "Remarks", "Latitude", "Longitude"]
    for col in target_order:
        if col in current_cols:
            current_cols.remove(col)
            
    # Find the index of the Date columns to slice the list
    insert_idx = len(current_cols) # Default to the end
    if DATE_APPROVED_COL in current_cols:
        insert_idx = current_cols.index(DATE_APPROVED_COL) + 1
        
    # If the End Date column exists, make sure we insert *after* it
    if DATE_END_COL in current_cols:
        insert_idx = max(insert_idx, current_cols.index(DATE_END_COL) + 1)
        
    # Reassemble the final column order
    final_cols = current_cols[:insert_idx] + target_order + current_cols[insert_idx:]
    
    # Ensure we only include columns that actually exist in the df to prevent KeyError
    final_cols = [c for c in final_cols if c in export_df.columns]
    
    return export_df[final_cols]


def render_filters(clean_df, pin_style, enable_clustering, show_legend, scale_by_funding):
    """Render the sidebar filter controls and return the filtered DataFrame."""
    st.sidebar.header("🔍 Filter Dashboard")

    available_divisions = clean_df["Division"].unique().tolist()
    selected_divisions = st.sidebar.multiselect(
        "Select Project(s):", available_divisions, default=available_divisions
    )

    available_statuses = clean_df["Map_Status"].unique().tolist()
    selected_statuses = st.sidebar.multiselect(
        "Select Project Status:", available_statuses, default=available_statuses
    )

    filtered_df = clean_df[
        clean_df["Division"].isin(selected_divisions)
        & clean_df["Map_Status"].isin(selected_statuses)
    ]

    if DATE_APPROVED_COL in clean_df.columns:
        approval_dates = pd.to_datetime(clean_df[DATE_APPROVED_COL], errors="coerce")
        valid_dates = approval_dates.dropna()
        
        if not valid_dates.empty:
            min_date = valid_dates.min().date()
            max_date = valid_dates.max().date()
        else:
            min_date = pd.Timestamp("2010-01-01").date()
            max_date = pd.Timestamp.today().date()
        
        date_range = st.sidebar.date_input(
            "Filter by Approval Date:",
            value=(min_date, max_date),
            min_value=min_date,
            max_value=max_date,
        )
        
        include_missing_dates = st.sidebar.checkbox("Include projects with missing dates", value=True)

        if isinstance(date_range, tuple):
            if len(date_range) == 2:
                start_dt = pd.to_datetime(date_range[0])
                end_dt = pd.to_datetime(date_range[1])
                filtered_approval_dates = pd.to_datetime(filtered_df[DATE_APPROVED_COL], errors="coerce", format="mixed")
                
                date_mask = (filtered_approval_dates >= start_dt) & (filtered_approval_dates <= end_dt)
                if include_missing_dates:
                    date_mask = date_mask | filtered_approval_dates.isna()
                    
                filtered_df = filtered_df[date_mask]
            elif len(date_range) == 1:
                start_dt = pd.to_datetime(date_range[0])
                filtered_approval_dates = pd.to_datetime(filtered_df[DATE_APPROVED_COL], errors="coerce", format="mixed")
                
                date_mask = (filtered_approval_dates >= start_dt)
                if include_missing_dates:
                    date_mask = date_mask | filtered_approval_dates.isna()
                    
                filtered_df = filtered_df[date_mask]

    if EFFECTIVE_FUNDING_COL in clean_df.columns:
        funding_numeric = pd.to_numeric(clean_df[EFFECTIVE_FUNDING_COL], errors="coerce")
        valid_funding = funding_numeric.dropna()
        
        if not valid_funding.empty:
            min_budget = int(valid_funding.min()) - 1
            max_budget = int(valid_funding.max()) + 1
            
            if max_budget > min_budget:
                budget_range = st.sidebar.slider(
                    "Filter by Budget (₱):",
                    min_value=min_budget,
                    max_value=max_budget,
                    value=(min_budget, max_budget),
                    step=1000,
                )
                
                include_missing_budget = st.sidebar.checkbox("Include projects with missing budget", value=True)
                
                filtered_funding = pd.to_numeric(filtered_df[EFFECTIVE_FUNDING_COL], errors="coerce")
                
                budget_mask = (filtered_funding >= budget_range[0]) & (filtered_funding <= budget_range[1])
                if include_missing_budget:
                    budget_mask = budget_mask | filtered_funding.isna()
                    
                filtered_df = filtered_df[budget_mask]

    st.sidebar.markdown("---")
    st.sidebar.metric(label="Projects Displayed", value=len(filtered_df))

    # --- NEW: Generate Presentation-Ready Data ---
    presentation_df = format_export_data(filtered_df)

    if st.sidebar.button("📄 View Raw Data Table", width="stretch"):
        show_raw_data(presentation_df) # Pass the clean data to the modal

    # --- Task 4.2: Filtered CSV Export ---
    csv_data = presentation_df.to_csv(index=False).encode('utf-8')
    
    st.sidebar.download_button(
        label="📥 Download Filtered Data (CSV)",
        data=csv_data,
        file_name="DOST_Davao_Filtered_Projects.csv",
        mime="text/csv",
        width="stretch" 
    )

    if st.sidebar.button("🖼️ Map Export Options", width="stretch"):
        st.session_state.pop("export_preview_bytes", None) # clear old preview
        show_export_dialog(filtered_df, pin_style, enable_clustering, show_legend, scale_by_funding)

    return filtered_df


def render_kpi_scorecards(filtered_df):
    """Render the five-column KPI summary above the map."""
    kpi1, kpi2, kpi3, kpi4, kpi5 = st.columns(5)

    with kpi1:
        st.metric(label="Total Projects", value=len(filtered_df))

    with kpi2:
        ongoing_count = len(filtered_df[filtered_df["Map_Status"] == "Ongoing"])
        st.metric(label="🚀 Ongoing", value=ongoing_count)

    with kpi3:
        completed_count = len(filtered_df[filtered_df["Map_Status"] == "Completed"])
        st.metric(label="✅ Completed", value=completed_count)

    with kpi4:
        terminated_count = len(filtered_df[filtered_df["Map_Status"] == "Terminated"])
        st.metric(label="🛑 Terminated", value=terminated_count)

    with kpi5:
        if EFFECTIVE_FUNDING_COL in filtered_df.columns:
            total_funding = pd.to_numeric(filtered_df[EFFECTIVE_FUNDING_COL], errors="coerce").sum()
            formatted_funding = format_currency(total_funding)
        else:
            formatted_funding = "N/A"

        st.metric(label="💰 Total Funding", value=formatted_funding)


def handle_map_click(map_data, filtered_df):
    """
    Open the project detail modal for whatever was clicked on the map.

    st_folium re-returns the same last_object_clicked value on every rerun
    until a new marker is clicked, so the click is tracked in session_state
    to avoid re-opening the dialog on unrelated reruns (e.g. a sidebar
    button press, which would otherwise collide with an already-open
    dialog and raise a StreamlitAPIException).
    """
    clicked = map_data.get("last_object_clicked")
    if not clicked:
        return

    click_key = (round(clicked["lat"], 6), round(clicked["lng"], 6))
    if click_key == st.session_state.get("last_handled_click"):
        return

    st.session_state["last_handled_click"] = click_key

    clicked_projects = filtered_df[
        (abs(filtered_df["Lat"] - clicked["lat"]) < CLICK_TOLERANCE_DEGREES)
        & (abs(filtered_df["Long"] - clicked["lng"]) < CLICK_TOLERANCE_DEGREES)
    ]

    if not clicked_projects.empty:
        show_project_details(clicked_projects)

# ==========================================
# 9. AI ASSISTANT & Q&A LOGIC (LOCAL PRIVACY VIA OLLAMA)
# ==========================================

def get_ollama_url():
    """
    Retrieve the active Ngrok tunnel link from Streamlit secrets.

    Returns the string URL if configured, otherwise returns None.
    This replaces the previous get_gemini_client method to ensure
    absolute data privacy locally on your hardware.
    """
    if "OLLAMA_URL" in st.secrets:
        return st.secrets["OLLAMA_URL"]
    return None


def generate_executive_summary(ollama_url, df):
    """Generate a stakeholder-ready executive summary using the local Phi-3 model."""
    clean_df = format_export_data(df)
    
    # Generate high-level summary statistics to avoid context-overflow
    total_projects = len(clean_df)
    status_counts = clean_df['Status'].value_counts().to_dict() if 'Status' in clean_df.columns else {}
    
    # Prefer the canonical effective-funding figure (Revised if present, else
    # Original); only fall back to keyword sniffing if it's missing.
    funding_col = "Funding (Effective)" if "Funding (Effective)" in clean_df.columns else next(
        (c for c in clean_df.columns if 'fund' in c.lower() or 'cost' in c.lower() or 'amount' in c.lower()), None
    )
    total_funding = 0
    if funding_col:
        total_funding = pd.to_numeric(clean_df[funding_col].astype(str).str.replace(r'[^\d.]', '', regex=True), errors='coerce').sum()

    div_col = next((c for c in clean_df.columns if 'div' in c.lower() or 'unit' in c.lower()), None)
    div_counts = clean_df[div_col].value_counts().to_dict() if div_col else {}

    summary_text = f"""
    Total Projects: {total_projects}
    Projects by Status: {status_counts}
    Total Funding: Php {total_funding:,.2f}
    Projects by Division: {div_counts}
    """

    # Limit rows to avoid huge prompt payloads that slow down local models
    MAX_ROWS = 50
    sample_df = clean_df.head(MAX_ROWS)
    csv_data = sample_df.to_csv(index=False)
    
    row_notice = ""
    if len(clean_df) > MAX_ROWS:
        row_notice = f"(Note: The CSV data below has been truncated to the top {MAX_ROWS} rows out of {len(clean_df)} to improve AI speed. Use the Data Summary above for accurate totals.)"

    prompt = f"""
    Please write a 2-paragraph executive summary based ONLY on the following regional project data summary and CSV sample.
    Focus on highlighting key funding allocations, division focus, and overall project health (Ongoing vs Completed vs Terminated).
    Do not use external knowledge.

    Use markdown to format your response neatly using bullets, highlighting, and headers. 
    Do not use h1 (#) or h2 (##) headers, only use h4 (####) if you want to use a header.

    Data Summary (Use these for overall totals):
    {summary_text}

    CSV Data Sample {row_notice}:
    {csv_data}
    """

    payload = {"model": OLLAMA_MODEL, "prompt": prompt, "stream": False}

    # Crucial headers to bypass Ngrok free-tier interceptor page
    headers = {
        "Content-Type": "application/json",
        "ngrok-skip-browser-warning": "true",
    }

    try:
        response = requests.post(
            f"{ollama_url}/api/generate", json=payload, headers=headers, timeout=600
        )
        response.raise_for_status()
        return response.json().get("response", "No response generated.")
    except Exception as e:
        return f"⚠️ Error generating local summary: {str(e)}"


def ask_ai_about_data(ollama_url, df, user_query, chat_history):
    """Answer natural-language data questions using the local Phi-3 instance with conversation history."""
    clean_df = format_export_data(df)
    
    # Limit rows to avoid huge prompt payloads that slow down local models
    MAX_ROWS = 50
    sample_df = clean_df.head(MAX_ROWS)
    csv_data = sample_df.to_csv(index=False)
    
    row_notice = ""
    if len(clean_df) > MAX_ROWS:
        row_notice = f"(Note: The data has been truncated to the top {MAX_ROWS} rows out of {len(clean_df)} to improve AI speed. Summarize based on this sample.)"

    # Format historical turns for context mapping
    history_text = "\n".join(
        [f"{msg['role'].capitalize()}: {msg['content']}" for msg in chat_history]
    )

    prompt = f"""
    You are an AI data assistant for the DOST-Davao Regional Office.
    Here is the currently filtered DOST project data in CSV format {row_notice}:

    {csv_data}

    User Query: {user_query}

    Based ONLY on the provided CSV data and the conversation history, answer the user's query.
    Be concise and professional, avoid using emdashes and just be functional with your tone.

    Use markdown to format your response in a neat way using bullets and highlighting, do not use headers.
    """

    payload = {"model": OLLAMA_MODEL, "prompt": prompt, "stream": False}

    headers = {
        "Content-Type": "application/json",
        "ngrok-skip-browser-warning": "true",
    }

    try:
        response = requests.post(
            f"{ollama_url}/api/generate", json=payload, headers=headers, timeout=600
        )
        response.raise_for_status()
        return response.json().get("response", "No response generated.")
    except Exception as e:
        return f"⚠️ Error communicating with Local AI: {str(e)}"

# ==========================================
# 10. MAP EXPORT & BRANDING
# ==========================================
def capture_map_screenshot(folium_map):
    """Save folium map to HTML and capture a screenshot using headless Chrome."""
    options = Options()
    options.add_argument('--headless')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--window-size=1920,1080')
    options.add_argument('--hide-scrollbars')
    
    # Try to use webdriver-manager, fallback to basic Chrome if it fails
    try:
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=options)
    except Exception:
        driver = webdriver.Chrome(options=options)
        
    try:
        with tempfile.NamedTemporaryFile(suffix='.html', delete=False) as tmp:
            html_path = tmp.name
            
            # Hide scrollbars specifically for the screenshot and force full width to avoid gaps
            folium_map.get_root().header.add_child(folium.Element("<style>body, html { margin:0 !important; padding:0 !important; overflow: hidden !important; width: 100vw !important; height: 100vh !important; } .folium-map { width: 100vw !important; height: 100vh !important; }</style>"))
            folium_map.save(html_path)
            
        driver.get(f"file:///{html_path}")
        time.sleep(3) # Wait for map tiles and clusters to load
        
        png_data = driver.get_screenshot_as_png()
        return Image.open(io.BytesIO(png_data))
    finally:
        driver.quit()
        # NEW: Delete the temporary HTML file so the server disk doesn't fill up
        try:
            os.remove(html_path)
        except Exception:
            pass

def create_branded_export(map_image, selected_divisions, selected_statuses, title_text="DOST-Davao Regional Project Mapping Report"):
    """Combine the map screenshot with DOST branding and active filter details."""
    canvas_width = map_image.width
    canvas_height = map_image.height + 250
    
    # Create canvas
    canvas = Image.new('RGB', (canvas_width, canvas_height), '#222d32')
    draw = ImageDraw.Draw(canvas)
    
    # Paste Map at the bottom
    canvas.paste(map_image, (0, 250))
    
    text_x = 250
    # Try to load and paste logo
    try:
        logo = Image.open("assets/dost_davao_logo.png").convert("RGBA")
        # Resize logo to fit nicely in header (e.g. max height 150)
        ratio = 150 / logo.height
        new_width = int(logo.width * ratio)
        logo = logo.resize((new_width, 150))
        # Paste at left, centered vertically in the 250px header
        canvas.paste(logo, (50, 50), logo)
        
        # Shift text to the right of the logo with a 50px buffer
        text_x = 40 + new_width + 40
    except Exception:
        pass
        
    # Attempt to load a default font, fallback to default PIL font
    try:
        font_title = ImageFont.truetype("assets/Montserrat-Bold.ttf", 50)
        font_subtitle = ImageFont.truetype("assets/Montserrat-Regular.ttf", 36)
    except Exception:
        font_title = ImageFont.load_default()
        font_subtitle = ImageFont.load_default()
        
    # Text Placement
    div_str = "All" if len(selected_divisions) > 2 else ", ".join(selected_divisions)
    stat_str = "All" if len(selected_statuses) > 2 else ", ".join(selected_statuses)
    subtitle_text = f"Divisions: {div_str} | Status: {stat_str}"
    
    draw.text((text_x, 60), title_text, font=font_title, fill=(255, 255, 255))
    draw.text((text_x, 135), subtitle_text, font=font_subtitle, fill=(255, 255, 255))
    
    # Export to bytes
    img_byte_arr = io.BytesIO()
    canvas.save(img_byte_arr, format='PNG')
    img_byte_arr.seek(0)
    
    return img_byte_arr.getvalue()

# ==========================================
# 11. MAIN APPLICATION LOGIC
# ==========================================
if uploaded_file is not None:
    with st.spinner("Processing..."):
        clean_df = load_and_clean_data(uploaded_file, ingestion_mode, selected_sheet)

    if clean_df.empty:
        st.error(
            f"⚠️ No projects with valid coordinates were found in this file under the "
            f"**{data_format_label}** setting. The map only displays projects with "
            f"latitude and longitude information (or a Coordinates column). If this "
            f"file does not contain location data, try selecting a different worksheet "
            f"or a different Data Format in the sidebar."
        )
        st.stop()

    # Persist a working copy across reruns (keyed to this exact file/mode/
    # sheet combination) so geocoded coordinates survive button clicks and
    # other widget interactions instead of resetting every rerun.
    dataset_key = (
        getattr(uploaded_file, "name", None),
        getattr(uploaded_file, "size", None),
        ingestion_mode,
        selected_sheet,
    )
    if st.session_state.get("_dataset_key") != dataset_key:
        st.session_state["_dataset_key"] = dataset_key
        st.session_state["working_df"] = clean_df.copy()
        st.session_state["geocode_failures"] = []

    working_df = st.session_state["working_df"]
    missing_coords_count = int((working_df["Lat"].isna() | working_df["Long"].isna()).sum())
    st.sidebar.header("📍 Missing Coordinates")
    if missing_coords_count == 0:
        st.sidebar.caption("Every project has coordinates.")
    else:
        st.sidebar.caption(
            f"{missing_coords_count} project(s) have no Lat/Long yet and won't "
            "appear on the map or in KPIs until resolved."
        )
        mapbox_token = get_mapbox_token()
        if not mapbox_token:
            st.sidebar.warning(
                "Mapbox isn't configured yet. Add `MAPBOX_API_KEY` to "
                "`.streamlit/secrets.toml` to enable this."
            )
        elif st.sidebar.button(f"📍 Resolve Missing Coordinates ({missing_coords_count})"):
            updated_df, failures = resolve_missing_coordinates(working_df, mapbox_token)
            st.session_state["working_df"] = updated_df
            st.session_state["geocode_failures"] = failures
            working_df = updated_df

            resolved_count = missing_coords_count - len(failures)
            if failures:
                st.sidebar.warning(f"✅ Resolved {resolved_count}. ⚠️ {len(failures)} need manual review.")
            else:
                st.sidebar.success(f"✅ All {resolved_count} coordinate(s) resolved!")

    if st.session_state.get("geocode_failures"):
        with st.sidebar.expander(f"⚠️ Needs Manual Review ({len(st.session_state['geocode_failures'])})"):
            st.dataframe(
                pd.DataFrame(st.session_state["geocode_failures"]),
                width="stretch",
                hide_index=True,
            )
            st.caption(
                "These couldn't be automatically geocoded. Add coordinates "
                "manually in the source file, or refine the Location / "
                "Implementing Agency text and re-run."
            )

    mappable_df = working_df.dropna(subset=["Lat", "Long"])


    if mappable_df.empty:
        st.error(
            "⚠️ No projects with usable coordinates yet. Use "
            "**📍 Resolve Missing Coordinates** in the sidebar, or check that "
            "this file/sheet actually has location data."
        )
        st.stop()


    # Create a map options section in the sidebar
    st.sidebar.divider()
    st.sidebar.header("🗺️ Map Options")
    pin_style = st.sidebar.radio(
        "Map Pin Style:",
        options=["Abbreviation Badge", "Standard Color Pin"],
        index=0
    )

    # Enable or disable clustering
    enable_clustering = st.sidebar.checkbox("Enable Marker Clustering", value=True)
    show_legend = st.sidebar.checkbox("Show Map Legend", value=True)
    scale_by_funding = st.sidebar.checkbox("Scale Markers by Funding", value=False)
    
    st.sidebar.divider()

    filtered_df = render_filters(mappable_df, pin_style, enable_clustering, show_legend, scale_by_funding)
    render_kpi_scorecards(filtered_df)

    project_map = create_map(filtered_df, pin_style, enable_clustering, show_legend, scale_by_funding)
    map_key = f"main_map_{show_legend}_{enable_clustering}_{pin_style}_{scale_by_funding}"
    map_data = st_folium(
        project_map,
        use_container_width=True,
        height=460,
        returned_objects=["last_object_clicked"],
        key=map_key,
    )

    handle_map_click(map_data, filtered_df)

    col1, col2 = st.columns([0.85, 0.15])
    with col2:
        if st.button("🔍 Maximize Map", use_container_width=True):
            show_fullscreen_map(filtered_df, pin_style, enable_clustering, show_legend, scale_by_funding)

    # (Removed original high-res export block from main body)

    # ==========================================
    # TASKS 5.2, 5.3, 5.4: AI CHAT INTERFACE
    # ==========================================
    st.markdown("---")
    
    # Use columns to put the Clear Chat button neatly on the right side
    head_col1, head_col2 = st.columns([4, 1])
    with head_col1:
        st.markdown("### 🗾 DOST-Davao Project Mapper AI")
    
    ollama_url = get_ollama_url()
    if ollama_url:
        with head_col2:
            if st.button("🗑️ Clear Chat", width="stretch"):
                st.session_state.chat_messages = [
                    {"role": "assistant", "content": "Hello! I am your DOST-Davao Project Mapper Data Assistant. How can I help you analyze the map data today?"}
                ]
                st.rerun()

        if st.button("📊 Generate Executive Summary for Current Map", width="stretch"):
            with st.spinner("Drafting executive summary..."):
                summary = generate_executive_summary(ollama_url, filtered_df)
                st.info(summary)
        
        st.caption("Ask natural language questions about the projects currently visible on the map.")
        
        # Initialize chat history in Streamlit Session State
        if "chat_messages" not in st.session_state:
            st.session_state.chat_messages = [
                {"role": "assistant", "content": "Hello! I am your DOST-Davao Project Mapper Data Assistant. How can I help you analyze the map data today?"}
            ]

        # Render existing chat messages
        for message in st.session_state.chat_messages:
            with st.chat_message(message["role"]):
                st.markdown(message["content"])

        # Chat input and AI response
        phrase_list = ["What is the total funding for ongoing " + random.choice(DATA_SHEETS) + " projects?",
                        "What are the date of completion for ongoing " + random.choice(DATA_SHEETS) + " projects?",
                        "What are the date of approval for ongoing " + random.choice(DATA_SHEETS) + " projects?",
                        "What are the funding amounts for ongoing " + random.choice(DATA_SHEETS) + " projects?",
                        "What are the project titles for " + random.choice(DATA_SHEETS) + " projects?",
                        "What are the project abbreviation for " + random.choice(DATA_SHEETS) + " projects?",
                        "What are the project division for " + random.choice(DATA_SHEETS) + " projects?",
                        "What are the project coordinates for " + random.choice(DATA_SHEETS) + " projects?",
                        ]

        # Initialize random placeholder in session state if it doesn't exist
        if "chat_placeholder" not in st.session_state:
            st.session_state.chat_placeholder = f"Ask a question about the current visible on the map. Ex. {random.choice(phrase_list)}"

        # Use the static placeholder from session state
        if prompt := st.chat_input(st.session_state.chat_placeholder):
            
            # 1. Display user message
            with st.chat_message("user"):
                st.markdown(prompt)
                
            # 2. Get the AI response
            with st.chat_message("assistant"):
                with st.spinner("Analyzing map data..."):
                    ai_response = ask_ai_about_data(ollama_url, filtered_df, prompt, st.session_state.chat_messages)
                    st.markdown(ai_response)
            
            # 3. Append both to session state
            st.session_state.chat_messages.append({"role": "user", "content": prompt})
            st.session_state.chat_messages.append({"role": "assistant", "content": ai_response})

    else:
        st.error("🔑 Ollama URL not found. Please add OLLAMA_URL to your `.streamlit/secrets.toml` file.")

else:
    st.info("Upload the raw Excel file in the sidebar to begin.")
